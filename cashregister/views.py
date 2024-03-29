import locale
from typing import Any
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import *
from cashregister.forms import CashRegisterForm, MovementForm, CurrencyForm
from django.urls import reverse_lazy
from django.db.models import Q
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse

#Importaciones de la app
from branches.utils import set_branch_session
from core.mixins import CustomUserPassesTestMixin, LoginRequiredMixin
from cashregister.models import CashRegister, CashRegisterDetail, PaymentType, Movement
from cashregister.forms import CloseCashRegisterForm, CashRegisterDetailFormSet
from sales.models import Payment
from cashregister.utils import obtener_nombres_de_campos
from django.views.generic import DetailView
from django.utils import timezone




# Create your views here.
class CashRegisterCreateView(LoginRequiredMixin, FormView):
    template_name = 'cashregister/cashregister_create_page.html'
    form_class = CashRegisterForm
    success_url = reverse_lazy('cashregister_app:cashregister_view')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        branch_actualy = set_branch_session(self.request)

        cashregister = CashRegister.objects.filter(branch=branch_actualy, is_close=False)

        context['cashregister'] = cashregister
        context['currency_form'] = CurrencyForm
        return context
    
    def form_valid(self, form):
        # branch = self.request.user.branch
        initial_balance = form.cleaned_data['initial_balance']
        final_balance = initial_balance
        user_made = self.request.user
        currency = form.cleaned_data['currency']
        
        branch_actualy = set_branch_session(self.request)
        try:
            CashRegister.objects.create_cash_register(initial_balance, branch_actualy, user_made, currency, final_balance)
        except Exception as e:
            messages.error(self.request, 'Existe al intentar abrir una caja. Consulte al administrador del sistema por este mensaje')
            return super().form_invalid(form)

        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Existe un error en el formulario. Consulte al administrador del sistema por este mensaje')
        return super().form_invalid(form)

class CashRegisterView(LoginRequiredMixin, TemplateView):
    template_name = 'cashregister/cashregister_page.html'
    model = CashRegister
    form_class = CloseCashRegisterForm
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Aquí se recupera la caja de la sucursal correspondiente al usuario logueado
        branch_actualy = set_branch_session(self.request)

        cashregister = CashRegister.objects.filter(branch=branch_actualy, is_close=False).last()

        context['cashregister'] = cashregister
        return context


class CashRegisterListView(LoginRequiredMixin, ListView):
    template_name = 'cashregister/cashregister_list_page.html'
    model = CashRegister
    
    def get_context_data(self, **kwargs: Any):
        context = super().get_context_data(**kwargs)
        # Aquí se recupera la caja de la sucursal correspondiente al usuario logueado
        
        branch_actualy = set_branch_session(self.request)
        try:
            cashregisters = CashRegister.objects.filter(branch=branch_actualy, deleted_at=None).order_by('-created_at')
        except CashRegister.DoesNotExist:
            cashregisters = None
            messages.error(self.request, 'No hay una caja registradora activa para esta sucursal')
        context['cashregisters'] = cashregisters
        context['table_column'] = obtener_nombres_de_campos(CashRegister,
            "id",
            "date_close", 
            "counted_balance", 
            "difference",
            "branch", 
            "deleted_at", 
            "created_at", 
            "updated_at", 
            "currency",
            )
        
        return context


class MovimientosCajaView(ListView):
    model = Movement
    template_name = 'cashregister/movements_cash_register_page.html'
    context_object_name = 'movements'
    
    def get_queryset(self):
        # Obtener el id del CashRegister de la URL
        cashregister_id = self.kwargs.get('cashregister_id')
        
        # Obtener el CashRegister correspondiente o retornar 404 si no existe
        cashregister = get_object_or_404(CashRegister, pk=cashregister_id)

        # Filtrar movimientos por el CashRegister y sin eliminar
        queryset = Movement.objects.filter(cash_register=cashregister, deleted_at=None).order_by('-created_at')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener el CashRegister de la URL
        cashregister_id = self.kwargs.get('cashregister_id')
        cashregister = get_object_or_404(CashRegister, pk=cashregister_id)

        context['cashregister'] = cashregister
        context['table_column_movement'] = obtener_nombres_de_campos(
            Movement,
            "description",
            "currency",
            "transaction",
            "id",
            "deleted_at",
            "created_at",
            "updated_at",
            "withdrawal_reason",
            "transaction",
            "cash_register",
        )

        return context

class CashRegisterDetailView(LoginRequiredMixin, DetailView):
    template_name = 'cashregister/cashregister_detail_page.html'
    model = CashRegister
    
    def get_context_data(self, **kwargs: Any):
        context = super().get_context_data(**kwargs)
        
        cashregister = self.get_object()
        movements = Movement.objects.filter(cash_register=cashregister, deleted_at=None).order_by('-created_at')[:5]
        
        context['cashregister'] = cashregister
        context['movements'] = movements
        context['table_column_movement'] = obtener_nombres_de_campos(Movement, 
            "description", 
            "currency", 
            "transaction",
            "id", 
            "deleted_at", 
            "created_at", 
            "updated_at", 
            "withdrawal_reason",
            "transaction",
            "cash_register",
        )
        
        # Obtener archering_data
        archering_data = CashRegister.objects.get_archering_data(self.get_object())
        # Invertir el orden de archering_data
        archering_data = dict(reversed(list(archering_data.items())))
        context['archering_data'] = archering_data
        return context

#Esta view aun esta sin uso
class CashRegisterUpdateView(CustomUserPassesTestMixin, DetailView):
    template_name = 'cashregister/cashregister_update_page.html'
    model = CashRegister

#Esta view aun esta sin uso
class CashRegisterDeleteView(CustomUserPassesTestMixin, DeleteView):
    template_name = 'cashregister/cashregister_delete_page.html'
    model = CashRegister
    success_url = reverse_lazy('cashregister_app:cashregister_list_view')


class CashRegisterCloseView(LoginRequiredMixin, FormView):
    template_name = 'cashregister/cashregister_close_page.html'
    form_class = CloseCashRegisterForm
    success_url = reverse_lazy('cashregister_app:cashregister_view')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        
        branch_actualy = set_branch_session(self.request)

        cashregister = CashRegister.objects.filter(is_close=False, branch=branch_actualy).last()
        if not cashregister:
            messages.error(self.request, 'No hay una caja registradora activa para esta sucursal')

        context['cashregister'] = cashregister
        return context
    
    def form_valid(self, form):
        # branch = self.request.user.branch
        branch_actualy = set_branch_session(self.request)
        observations = form.cleaned_data['observations']
        cashregister = CashRegister.objects.filter(is_close=False, branch=branch_actualy).last()
        cashregister.observations = observations
        cashregister.is_close = True
        cashregister.save()

        # Redirige a la lista de cajas después de cerrar la caja
        messages.success(self.request, 'La caja se cerró correctamente.')
        return redirect('cashregister_app:cashregister_list_view')
    
    def form_invalid(self, form):
        messages.error(self.request, 'Existe un error en el formulario. Consulte al administrador del sistema por este mensaje')
        return super().form_invalid(form)


class CashRegisterArching(LoginRequiredMixin, View):
    template_name = 'cashregister/cashregister_arching_page.html'
    
    def get(self, request, *args, **kwargs):
        
        branch_actualy = set_branch_session(self.request)

        cashregister = CashRegister.objects.filter(is_close=False, branch= branch_actualy).last() #pasar contexto de cashregiter
        if not cashregister:
            messages.error(request, 'No hay una caja registradora activa para esta sucursal')
            return redirect('cashregister_app:cashregister_create_view')
        
        initial_data = [
            {
                'type_method': method, # tipo de metodo de pago
                'registered_amount': CashRegisterDetail.objects.registered_amount_for_type_method(method, Movement, cashregister) # calculo de monto registrado por el metodo de pago elegido
            } for method in PaymentType.objects.exclude(name='Cuenta Corriente')]
        # Hay que buscar una alternativa a esta linea de codigo
        # Porque todo deja de funcionar si se saca initial=initial_data
        formset = CashRegisterDetailFormSet(initial=initial_data)
        context = {'cashregister': cashregister, 'formset': formset}
        
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        branch_actualy = set_branch_session(self.request)

        cashregister = CashRegister.objects.filter(is_close=False, branch= branch_actualy).last() #pasar contexto de cashregiter
        if not cashregister:
            messages.error(request, 'No hay una caja registradora activa para esta sucursal')

        formset = CashRegisterDetailFormSet(request.POST)
        final_data = [
            {
                'type_method': method,
                'registered_amount': CashRegisterDetail.objects.registered_amount_for_type_method(method, Movement, cashregister)
            }
            for method in PaymentType.objects.all()]
        if formset.is_valid():
            
            branch_actualy = set_branch_session(self.request)
            cashregister = CashRegister.objects.filter(is_close=False, branch=branch_actualy).last()
            
            for form, data in zip(formset, final_data): # Esto esta provisorio, machea cada metodo con un form del formset, pero si se cambian de orden se rompe todo
                if form.is_valid():
                    instance = form.save(commit=False)
                    instance.user_made = request.user
                    instance.cash_register = cashregister
                    instance.type_method = data['type_method']
                    instance.registered_amount = abs(data['registered_amount'])
                    instance.counted_amount = abs(instance.counted_amount) if instance.counted_amount is not None else 0
                    instance.difference = instance.counted_amount - instance.registered_amount
                    instance.save()
            # Se comento esto porque ahora no se cierra la caja sino que se registra una arqueo nada mas 
            # cashregister.is_close = True
            # cashregister.save()
            

            ########### HARDCODEO MORTAL PARA CUENTA CORREINTE ##############
            from django.db.models import Sum
            registered_amount = Payment.objects.filter(customer__user_made__branch=branch_actualy, payment_method__name='Cuenta Corriente', created_at__date=timezone.now().date()).aggregate(amount=Sum('sale__total'))['amount'] or 0
            CashRegisterDetail.objects.create(
                user_made = request.user,
                cash_register = cashregister,
                type_method = PaymentType.objects.get(name='Cuenta Corriente'),
                registered_amount = abs(registered_amount),
                counted_amount = abs(registered_amount),
                difference = 0
            )

        else:
            messages.error(request, 'Existe un error en el formulario. Consulte al administrador del sistema por este mensaje')
            return render(request, self.template_name, {'formset': formset})
        return redirect('cashregister_app:cashregister_view')


class CloseTicketCashRegister(LoginRequiredMixin, View):
    template_name = 'cashregister/components/ticket_close_cashregister.html'
    def get(self, request, *args, **kwargs):
        
        context = {}
        try:
            #recuperamos la caja que viene por argumento en la peticion get
            cashregister = CashRegister.objects.get(pk=kwargs['pk'])
        except CashRegister.DoesNotExist:
            cashregister = None
            messages.error(self.request, 'No hay una caja registradora activa para esta sucursal')
        context['cashregister'] = cashregister
        context['movements'] = CashRegister.objects.get_movements_data(cashregister)
        
        return render(request, self.template_name, context)


#Falta corregir esta funcion y pasarla a una clase
def archingTicket(request, pk, archiv_pos):
    total = 0
    
    cashregister = CashRegister.objects.get(pk=pk)
    archering_data = CashRegister.objects.get_archering_data(cashregister)
    for dicc in next(iter(archering_data[int(archiv_pos)].items()))[1]:
        total += dicc['counted_amount']
    total = int(total)


    # locale.setlocale(locale.LC_TIME, 'es_ES.utf8')
    format = "%A, %d de %B de %Y"
    
    created_at = next(iter(archering_data[int(archiv_pos)].items()))[0]
    sale_date_str = created_at.strftime(format)
    context = {
        'archering_data': archering_data[int(archiv_pos)],
        'total': total,
        'date': sale_date_str,
        'time': created_at.time(),
        }

    return render(request, 'cashregister/components/ticket_arching.html', context)


class MovementsView(LoginRequiredMixin, TemplateView):
    template_name = 'cashregister/movements_page.html'
    paginate_by = 50
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Aquí se recupera la caja de la sucursal correspondiente al usuario logueado
        
        branch_actualy = set_branch_session(self.request)

        cashregisters = CashRegister.objects.filter(branch=branch_actualy, deleted_at=None)

        #Tener en cuenta que cuando se hace una consulta por filtro anula lo de deleted_at y hay que especificarlo de nuevo
        movements = Movement.objects.filter(cash_register__in=cashregisters, deleted_at=None).order_by('-created_at')[:self.paginate_by]
        context['movements'] = movements
        context['table_column'] = obtener_nombres_de_campos(Movement, 
            "description", 
            "currency", 
            "transaction",
            "id", 
            "deleted_at", 
            "created_at", 
            "updated_at", 
            "withdrawal_reason",
            "transaction",
            "cash_register",
            "payment_method"
            )
        
        return context


class MovementsCreateView(LoginRequiredMixin, FormView):
    template_name = 'cashregister/movements_create_page.html'
    model = Movement
    form_class = MovementForm
    success_url = reverse_lazy('cashregister_app:movements_view')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        branch_actualy = set_branch_session(self.request)

        cashregister = CashRegister.objects.filter(branch=branch_actualy, is_close=False).last()
        if not cashregister:
            messages.error(self.request, 'No hay una caja registradora activa para esta sucursal')
        
        context['cashregister'] = cashregister
        
        return context
    
    def form_valid(self, form):
        branch_actualy = set_branch_session(self.request)
        cash_register = CashRegister.objects.filter(branch=branch_actualy, is_close=False).last()

        try:
            Movement.objects.update_balance(cash_register, form.instance.amount, form.instance.type_operation)
            form.instance.amount = abs(form.instance.amount)
            form.instance.date_movement = timezone.now()
            form.instance.cash_register = cash_register
            form.instance.currency = cash_register.currency
            form.instance.user_made = self.request.user
            form.save()
            
        except ValueError as e:
            messages.error(self.request, 'Ha ocurrido un error. No hay fondos suficientes para realizar la operacion')
            return redirect('cashregister_app:movements_create_view')

        messages.success(self.request, 'El movimiento fue creado con éxito')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        
        messages.error(self.request, 'Existe un error en el formulario. Consulte al administrador del sistema por este mensaje')
        
        return super().form_invalid(form)


class MovementsDeleteView(CustomUserPassesTestMixin, DeleteView):
    template_name = 'cashregister/movements_delete_page.html'
    model = Movement
    success_url = reverse_lazy('cashregister_app:movements_view')
    
    def delete(self, request, *args, **kwargs):
        
        messages.success(self.request, 'El movimiento fue eliminado con éxito')
        
        return super().delete(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = "Eliminar movimiento"
        context['form_description'] = "¿Está seguro que desea eliminar este movimiento?"
        return context
    
    def post(self, request, *args, **kwargs):
        messages.success(self.request, 'El movimiento fue eliminado con éxito')
        return super().post(request, *args, **kwargs)
    

class MovementsUpdateView(CustomUserPassesTestMixin, UpdateView):
    template_name = 'cashregister/movements_create_page.html'
    model = Movement
    form_class = MovementForm
    success_url = reverse_lazy('cashregister_app:movements_view')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        
        branch_actualy = set_branch_session(self.request)

        cashregister = CashRegister.objects.filter(branch=branch_actualy, is_close=False).last()
        if not cashregister:
            messages.error(self.request, 'No hay una caja registradora activa para esta sucursal')
        context['cashregister'] = cashregister
        return context
    
    def form_valid(self, form):
        cashregister = form.instance.cash_register # recuperamos la caja de la instancia del movimiento a actualizar
        movement = self.get_object() # este es el movimiento actual antes de ser modificado
        form.instance.amount = abs(form.instance.amount)
        
        if cashregister.is_close:
            messages.error(self.request, 'No se puede actualizar un movimiento en una caja cerrada')
            return redirect('cashregister_app:movements_view')
        
        try:
            #Luego esto se puede refactorizar porque al pasarle el movimiento ya tengo acceso a la caja de la instancia del movimiento a actualizar
            Movement.objects.update_balance_reverse(cashregister, form.instance.amount, form.instance.type_operation, movement)
            messages.success(self.request, 'El movimiento fue actualizado con éxito')
        except ValueError as e:
            messages.success(self.request, 'Se pudo realizar la reverza del movimento')
            messages.error(self.request, 'Ha ocurrido un error. No hay fondos suficientes para realizar la operacion')
            
            #recordando que tengo en mi managers  except ValueError as e:
            # Si ocurre un error, significa que estoy tratando de retirar un monto que no tengo, por lo que debo eliminar el movimiento
            # current_movement.delete()
            # e.args.__add__(("Movement deleted",))
            # raise e
            
            return redirect('cashregister_app:movements_create_view')
        
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Existe un error en el formulario. Consulte al administrador del sistema por este mensaje')
        return super().form_invalid(form)


class MovementsDetailView(LoginRequiredMixin, DetailView):
    template_name = 'cashregister/movements_detail_page.html'
    model = Movement
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #por ahora no se hizo nada ...
        return context


class CurrencyView(TemplateView):
    pass


class CurrencyCreateView(CustomUserPassesTestMixin, FormView):
    """
    Crear una catogoria nueva para el producto
    """
    form_class = CurrencyForm
    template_name = 'cashregister/currency_create_page.html'
    success_url = reverse_lazy('cashregister_app:currency_list')

    def form_valid(self, form):
        currency = form.save(commit=False)
        currency.user_made = self.request.user
        currency.save()

        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # Para saber si es una peticion AJAX
            new_currency_data = {
                'id': currency.id,
                'name': currency.name,
                'symbol': currency.symbol,
                'code': currency.code,
            }
            # Si es una solicitud AJAX, devuelve una respuesta JSON
            return JsonResponse({'status': 'success', 'new_currency': new_currency_data})
        else:
            # Si no es una solicitud AJAX, llama al método form_valid del padre para el comportamiento predeterminado
            return super().form_valid(form)


#------- VISTAS BASADAS EN FUNCIONES PARA PETICIONES AJAX -------#

################ SEARCH MOVEMENTS AJAX ################

def ajax_search_movements(request):
    # branch = request.user.branch

    branch_actualy = set_branch_session(request)

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        # Obtener el valor de search_term de la solicitud
        search_term = request.GET.get('search_term', '')
        search_terms = search_term.split()
        if not search_term:
            # En caso de que search_term esté vacío, muestra la cantidad de empleados por defecto
            paginate_by = MovementsView().paginate_by
            movements = Movement.objects.all().filter(deleted_at = None, cash_register__branch=branch_actualy)[:paginate_by]
        else:
            # Usando Q por todos los campos existentes en la tabla
            query_conditions = Q()
            for term in search_terms:
                query_conditions |= (
                    Q(amount__icontains=term) |
                    Q(date_movement__icontains=term) |
                    Q(type_operation__icontains=term) |
                    Q(user_made__first_name__icontains=term) |
                    Q(user_made__last_name__icontains=term)
                )

            movements = Movement.objects.filter(deleted_at=None, cash_register__branch=branch_actualy).filter(query_conditions)[:40]

        data = [{
            'id': movement.id,
            'amount': movement.amount,
            'date_movement': movement.date_movement.strftime('%Y-%m-%d'),
            'type_operation': movement.type_operation,
            'user_made': str(movement.user_made),
            'is_staff': 1 if request.user.is_staff else 0
        } for movement in movements]
        # locale.setlocale(locale.LC_TIME, '')
        return JsonResponse({'data': data})
    return redirect(reverse_lazy("cashregister_app:movements_view"))
    

############################### EXPORTAR MOVIMIENTOS ###############################

def export_movements_list_to_excel(request):
    # Para la generacion de excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    
    branch_actualy = set_branch_session(request)
    
    list_movements = Movement.objects.filter(cash_register__branch=branch_actualy)

    if not list_movements:
        raise ValueError('No hay productos para exportar.') # modificar error
    
    # Crear un libro de trabajo de Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Definir estilos personalizados para los encabezados
    header_style = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0b1727', end_color='0b1727', fill_type='solid')

    # Definir los encabezados de las columnas
    exclude_fields_user = ["deleted_at", "created_at", "updated_at", "id", "user_made", "transaction", "withdrawal_reason"]
    headers = [campo[1] for campo in obtener_nombres_de_campos(Movement, *exclude_fields_user)]

    # Aplicar estilos a los encabezados y escribir los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_style
        cell.fill = header_fill

    # Modificar el ancho de la columna (ajustar segÃºn tus necesidades)
    #################################################
    try: 
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter
    #################################################
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 25

    # Agregar los datos de los empleados a la hoja de cÃ¡lculo
    for row_num, movement in enumerate(list_movements, 2):
        worksheet.cell(row=row_num, column=1, value=str(movement.created_at.strftime("%d/%m/%Y %H:%M")))
        worksheet.cell(row=row_num, column=2, value=str(movement.cash_register))
        worksheet.cell(row=row_num, column=3, value=str(movement.amount))
        worksheet.cell(row=row_num, column=4, value=movement.description)
        worksheet.cell(row=row_num, column=5, value=str(movement.currency))
        worksheet.cell(row=row_num, column=6, value=movement.type_operation)
        worksheet.cell(row=row_num, column=7, value=str(movement.payment_method))

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lista de movimientos - Sucursal %s.xlsx' %(branch_actualy.name)

    workbook.save(response)

    return response