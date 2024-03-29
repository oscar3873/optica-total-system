import locale
from django.template import loader
from django.db import transaction
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy
from django.views.generic import (DeleteView, UpdateView, DetailView, FormView, ListView)
from django.contrib import messages

from core.mixins import CustomUserPassesTestMixin
from branches.utils import set_branch_session
from cashregister.utils import create_in_movement, obtener_nombres_de_campos
from cashregister.models import CashRegister, Currency, Movement
from sales.models import Payment, Sale
from sales.forms import PaymentMethodForm, TypePaymentMethodForm
from sales.utils import set_movement, up_objetives
from django.utils import timezone

from .models import *
from .forms import *
from .utils import *


########################### CREATE ####################################
class ServiceOrderCreateView(LoginRequiredMixin, FormView):
    model = ServiceOrder
    form_class = ServiceOrderForm
    template_name = 'clients/service_order_form.html'
    success_url = reverse_lazy('clients_app:service_order_view')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ServiceOrderForm()
        context['customer'] =  Customer.objects.get(pk=self.kwargs['pk'])
        context['order_service'] = {
            'pupilar': InterpupillaryForm,
            'correction': CorrectionForm,
            'material': MaterialForm,
            'color': ColorForm,
            'cristal': CristalForm,
            'tratamiento': TratamientForm,
        }
        return context
    
    @transaction.atomic
    def form_valid(self, form):
        try:
            customer = Customer.objects.get(pk=self.kwargs.get('pk')) # pk corresponde a como se le pasa por la url.py <pk>
        except Customer.DoesNotExist:
            raise ValueError('El ID del cliente con nuestro registro')
        
        correction_form = CorrectionForm(self.request.POST)
        material_form = MaterialForm(self.request.POST)
        color_form = ColorForm(self.request.POST)
        cristal_form = CristalForm(self.request.POST)
        tratamiento_form = TratamientForm(self.request.POST)
        pupilar_form = InterpupillaryForm(self.request.POST)

        if (
            correction_form.is_valid() and
            material_form.is_valid() and 
            color_form.is_valid() and
            cristal_form.is_valid() and 
            tratamiento_form.is_valid() and 
            pupilar_form.is_valid()
            ):

            messages.success(self.request, 'Se ha registrado una nueva orden de servicio con exito.')

        if customer:
            return redirect('clients_app:customer_detail', pk=self.kwargs.get('pk'))
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Por favor, verifique los campos.')
        return super().form_invalid(form)
    

class CustomerCreateView(LoginRequiredMixin, FormView):
    form_class = CustomerForm
    template_name = 'clients/customer_form.html'
    success_url = reverse_lazy('clients_app:customer_view')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['h_insurance'] = HealthInsuranceForm
        return context

    @transaction.atomic
    def form_valid(self, form):
        if form.is_valid():
            user = self.request.user
            
            customer = form.save(commit=False)
            customer.user_made = self.request.user
            
            customer.credit_balance = 0
            
            branch_actualy = set_branch_session(self.request)

            customer.branch = branch_actualy
            customer.save()

            for insurance in form.cleaned_data['h_insurance']:
                Customer_HealthInsurance.objects.create(
                    h_insurance = insurance,
                    customer = customer,
                    user_made=user
                )
            if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # Para saber si es una peticion AJAX                
                data = {
                    'id': customer.id,
                    'first_name': customer.first_name,
                    'last_name': customer.last_name,
                    'phone_number': customer.phone_number,
                    'phone_code': customer.phone_code,
                    'dni': customer.dni,
                    'credit_balance': '$ %s' % customer.credit_balance if customer.has_credit_account else None,
                    'has_credit_account': customer.has_credit_account
                }
                return JsonResponse({'customer': data})
            
        messages.success(self.request, 'Se ha registrado un cliente con exito.')
        return super().form_valid(form)

    def form_invalid(self, form):
        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # Para saber si es una peticion AJAX  
            return JsonResponse({'error': 'Por favor, verifique los campos'})
        messages.error(self.request, 'Por favor, verifique los campos.')
        return super().form_invalid(form)


class HealthInsuranceCreateView(LoginRequiredMixin, FormView):
    form_class = HealthInsuranceForm
    template_name = 'clients/insurance_form.html'
    success_url = reverse_lazy('clients_app:insurance_view')

    def form_valid(self, form):
        insurance = form.save(commit=False)
        insurance.user_made = self.request.user
        insurance.save()

        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # AGREGA NUEVA OBRA SOCIAL DURANTE CREACION DE CLIENTE (PETICION FETCH)
            new_insurance_data = {
                'id': insurance.id,
                'name': insurance.name
            }
            return JsonResponse({'status': 'success', 'new_insurance': new_insurance_data})
        else:
            messages.success(self.request, 'Se ha registrado una obra social con exito.')
            return super().form_valid(form)
    
    def form_invalid(self, form):
        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # ERROR DE NUEVA OBRA SOCIAL DURANTE CREACION DE CLIENTE (PETICION FETCH)
            return JsonResponse({'error':'Por favor, verifique los campos'})
        else:
            messages.error(self.request, 'Por favor, verifique los campos.')
            return super().form_invalid(form)


########################### UPDATE  #########################

class ServiceOrderUpdateView(LoginRequiredMixin, UpdateView):
    model = ServiceOrder
    form_class = ServiceOrderForm  # Actualizar al formulario principal si es necesario
    template_name = 'clients/service_order_form.html'
    success_url = reverse_lazy('clients_app:service_order_view')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['customer'] =  self.get_object().customer
        context['order_service'] = {
            'correction': CorrectionForm(instance=self.object.correction),
            'material': MaterialForm(instance=self.object.material),
            'color': ColorForm(instance=self.object.color),
            'cristal': CristalForm(instance=self.object.type_cristal),
            'tratamiento': TratamientForm(instance=self.object.tratamient),
            'pupilar': InterpupillaryForm(instance=self.object.interpupillary),
        }
        return context

    
    def form_valid(self, form):
        try:
            customer = self.get_object().customer
        except Customer.DoesNotExist:
            raise ValueError('El ID del cliente con nuestro registro')
        
        named_formsets = self.get_context_data()['order_service']

        if form.is_valid():
            service_order = form.save(commit=False)
            for prefix, formset in named_formsets.items():
                instance = formset.instance
                new_formset = formset.__class__(self.request.POST, instance=instance)
                if new_formset.is_valid():
                    new_formset.save()
                # else:
            service_order.save()
        messages.success(self.request, 'Se ha acutalizado la orden de servicio con exito.')
        if customer:
            return redirect('clients_app:customer_detail', pk=customer.pk)

        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Por favor, verifique los campos.')
        return super().form_invalid(form)
    
    
class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'clients/customer_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['h_insurance'] = HealthInsuranceForm
        context['update'] = 1
        return context

    def form_valid(self, form):
        customer = form.save(commit=False)
        customer.user_made = self.request.user
        customer.save()

        form_in_out_insurances(form, customer, self.request.user)

        messages.success(self.request, 'Se ha acutalizado los datos del cliente con exito.')
        return redirect('clients_app:customer_detail', pk=self.get_object().pk)

    def form_invalid(self, form):
        messages.error(self.request, 'Por favor, verifique los campos.')
        return super().form_invalid(form)
    
    
class CustomerUpdateHealthInsurance(CustomUserPassesTestMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'clients/update_customer_hinsurance.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['h_insurance'] = HealthInsuranceForm
        return context

    def form_valid(self, form):
        customer = form.instance
        customer.user_made = self.request.user
        customer.save()

        form_in_out_insurances(form, customer, self.request.user)

        messages.success(self.request, 'Se ha acutalizado la obra social con exito.')
        return redirect('clients_app:customer_detail', pk=self.get_object().pk)

    def form_invalid(self, form):
        messages.error(self.request, 'Por favor, verifique los campos.')
        return super().form_invalid(form)
    
    
class HealthInsuranceUpdateView(CustomUserPassesTestMixin, UpdateView):
    model = HealthInsurance
    form_class = HealthInsuranceForm
    template_name = 'clients/hinsurance_update.html'
    success_url = reverse_lazy('clients_app:insurance_view')
    
    def form_valid(self, form):
        insurance = form.save(commit=False)
        insurance.user_made = self.request.user
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Por favor, verifique los campos.')
        return super().form_invalid(form)

########################## DETAIL #################################


class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'clients/customer_detail.html'
    context_object_name = 'customer'

    def get(self, request, *args, **kwargs):
        user = request.user
        branch = user.branch
        customer = self.get_object()

        if 'consumidor' in customer.first_name.lower() and customer.pk == 1:
            messages.error(request, 'Lo sentimos, no puedes ver este cliente.')
            return redirect('clients_app:customer_view')

        if not user.is_staff and not customer.branch == branch: 
            # El usuario no tiene permiso para ver este cliente (no corresponde a su Sucursal)
            messages.error(request, 'Lo sentimos, no puedes ver este cliente.')
            return redirect('clients_app:customer_view')
        return super().get(request, *args, **kwargs)

    def get_context_data(self,**kwargs):
        customer = self.get_object()
        context = super().get_context_data(**kwargs)

        service_order = Customer.objects.history(customer)
        context['service_orders'] = service_order
        context['pay_form'] = TypePaymentMethodForm
        context['payment_method_form'] = PaymentMethodForm

        context['table_column'] = obtener_nombres_de_campos(Sale,
            "id",
            "deleted_at", 
            "created_at", 
            "updated_at",
            "receipt",
            "customer",
            "discount",
            "refund_date",
            "discount_extra",
            "subtotal",
            "branch",
            "total"
            )
        context['sales_pendientes'] = customer.sales.filter(state="PENDIENTE")

        return context
    
    
class ServiceOrderDetailView(LoginRequiredMixin, DetailView):
    model = ServiceOrder
    template_name = 'clients/service_order_detail.html'
    context_object_name = 'service_order'


class HealthInsuranceDetailView(LoginRequiredMixin, DetailView):
    model = HealthInsurance
    template_name = 'clients/hinsuranse_detail.html'
    context_object_name = 'h_insurance'

############################## LISTING ###############################

class CustomerListView(LoginRequiredMixin, ListView):
    model = Customer
    template_name = 'clients/customer_page.html'
    context_object_name = 'customers'
    paginate_by = 50

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['table_column'] = obtener_nombres_de_campos(Customer,
            "id",
            "deleted_at", 
            "created_at", 
            "updated_at",
            "phone_code",
            "branch",
            "birth_date",
            "address",
            "email",
            "credit_balance",
            )
        return context

    def get_queryset(self):
        
        branch_actualy = set_branch_session(self.request)

        return Customer.objects.filter(branch=branch_actualy, deleted_at=None)


class ServiceOrderListView(LoginRequiredMixin, ListView):
    model = ServiceOrder
    template_name = 'clients/service_order_page.html'
    context_object_name = 'laboratory_orders'
    paginate_by = 50


class HealthInsuranceListView(LoginRequiredMixin, ListView):
    model = HealthInsurance
    template_name = 'clients/hinsuranse_page.html'
    context_object_name = 'h_insurances'
    paginate_by = 50


########################### DELETE ####################################

class ServiceOrderDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = ServiceOrder
    template_name = 'clients/service_order_delete.html'
    context_object_name = 'laboratory'
    success_url = reverse_lazy('clients_app:service_order_view')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context['customer'] = self.get_object().customer
        return context

    def get_success_url(self):
        try:
            customer = self.get_object().customer
            return reverse_lazy('clients_app:customer_detail', kwargs={'pk': customer.pk}) # en caso de tener pk de customer, entonces entre al delete por la vista de un cliente especifico
        except:
            return super().get_success_url() # Si entre al delete desde la lista general de pedidos de lab
        

class CustomerDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = Customer
    template_name = 'clients/customer_delete.html'
    success_url = reverse_lazy('clients_app:customer_view')

    def delete(self, request, *args, **kwargs):
        customer = self.get_object()

        if 'consumidor' in customer.first_name.lower() or customer.pk == 1:
            messages.error(request, 'Lo sentimos, no puedes eliminar este cliente.')
            return reverse_lazy('clients_app:customer_view')

        intermedia = Customer_HealthInsurance.objects.filter(customer=customer)
        for filas in intermedia:
            filas.delete()
        return super().delete(request, *args, **kwargs)
    

class HealthInsuranceDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = HealthInsurance
    template_name = 'clients/hinsurance_delete.html'
    success_url = reverse_lazy('clients_app:insurance_view')
    
    def delete(self, request, *args, **kwargs):
        h_insurance = self.get_object()
        intermedia = Customer_HealthInsurance.objects.filter(h_insurance=h_insurance)
        for filas in intermedia:
            filas.delete()
        return super().delete(request, *args, **kwargs)
    

####################### CREDIT ACCOUNT #######################

def open_credit_account(request, pk): # usar con   --->   <form method='post' acction="{% url 'clients_app:open_credit_account' customer.pk %}"> </form>
    customer = Customer.objects.get(pk=pk)
    
    if request.method == 'POST':
        if not customer.has_credit_account:
            customer.has_credit_account = True
            # customer.credit_balance = 0
            customer.save()
            messages.success(request, "Se ha abierto una cuenta Corriente con exito.")
        else:
            messages.error(request, "Este cliente ya tiene una Cuenta corriente abierta.")
    return redirect('clients_app:customer_detail', pk=customer.pk)


def pay_credits(request, pk):
    user = request.user
    customer = Customer.objects.get(pk=pk)
    payment = TypePaymentMethodForm(request.POST)        
    
    if payment.is_valid():
        total = payment.cleaned_data['amount']
        pay = payment.save(commit=False)

        branch_actualy = set_branch_session(request)

        cashregister = CashRegister.objects.filter(is_close=False, branch=branch_actualy).last()
        if not cashregister:
            messages.error(request, 'Antes de realizar una Venta, debe Abrir una Caja.')
            return redirect('cashregister_app:cashregister_view')

        mov = create_in_movement(branch_actualy, request.user, pay.payment_method.type_method, pay.description, total)
        
        pay.customer = customer
        pay.movement = mov
        pay.save()
        
        customer.credit_balance -= total
        customer.user_made = user
        customer.save()

        messages.success(request, "Se recibio el pago.")
        return redirect('clients_app:customer_detail', pk=customer.pk)
    else:
        messages.error(request, "Se produjo un error al realizar pago.")
        return redirect('clients_app:customer_detail', pk=customer.pk)


def close_credit_account(request, pk):
    user = request.user
    customer = Customer.objects.get(pk=pk)

    if not user.is_staff:
        messages.warning(request, "Solo administradores pueden cerrar cuenta corriente.")
        return redirect('clients_app:customer_detail', pk=customer.pk)

    if customer.credit_balance == 0:
        customer.has_credit_account = False
        customer.user_made = user
        customer.save()
        messages.warning(request, "Se cerró Cuenta Corriente del cliente.")
    else:
        messages.warning(request, "El cliente posee deuda y no se puede cerrar la cuenta corriente.")
    
    return redirect('clients_app:customer_detail', pk=customer.pk)


################ EXPORT SERVICE ORDERS #####################

def export_order_service_list_to_excel(request, pk):
    # Para la generacion de excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    customer = Customer.objects.get(id=pk)
    service_order_of_customer = customer.serviceorders.all()

    if not service_order_of_customer:
        raise ValueError('El cliente no cuenta con Ordenes de Servicio.') # modificar error
    
    # Crear un libro de trabajo de Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Definir estilos personalizados para los encabezados
    header_style = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0b1727', end_color='0b1727', fill_type='solid')

    status_style_red = Font(name='Arial', size=10, bold=True, color='FF0000')
    status_style_green = Font(name='Arial', size=10, bold=True, color='008000')


    # Definir los encabezados de las columnas
    exclude_fields_user = ["deleted_at", "updated_at", "correction", "material", "type_cristal", "color", "tratamient", "interpupillary", "customer"]
    headers = [campo[1] for campo in obtener_nombres_de_campos(ServiceOrder, *exclude_fields_user)]

    if 'created at' in headers:
        index = headers.index('created at')
        headers[index] = "Fecha de registro"

    # Aplicar estilos a los encabezados y escribir los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_style
        cell.fill = header_fill

    # Modificar el ancho de la columna (ajustar segÃºn tus necesidades)
    #################################################
    try: 
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter
    #################################################
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 25

    # Agregar los datos de los empleados a la hoja de cÃ¡lculo
    for row_num, serviceorder in enumerate(service_order_of_customer, 2):
        worksheet.cell(row=row_num, column=1, value=str(serviceorder.id))
        worksheet.cell(row=row_num, column=2, value=str(serviceorder.created_at.date()))
        worksheet.cell(row=row_num, column=3, value=str(serviceorder.user_made))
        worksheet.cell(row=row_num, column=4, value=serviceorder.diagnostic)
        worksheet.cell(row=row_num, column=5, value=serviceorder.armazon)
        worksheet.cell(row=row_num, column=6, value=serviceorder.observations)
        worksheet.cell(row=row_num, column=7, value=('No compleado' if not serviceorder.is_done else 'Completado'))

        worksheet.cell(row=row_num, column=7).font = (status_style_red if not serviceorder.is_done else status_style_green)

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s_%s Orden de servicio.xlsx' %(customer.last_name, customer.first_name)

    workbook.save(response)

    return response


def pay_service_order(request, pk):
    if request.method == 'POST':
        service = ServiceOrder.objects.get(id=pk)
        customer = service.customer
        sale = service.sale

        branch_actualy = set_branch_session(request)
        
        form = TypePaymentMethodForm(request.POST)
        if form.is_valid():
            amount = form.cleaned_data['amount']
            payment_method = form.cleaned_data['payment_method']
            mov = create_in_movement(branch_actualy, request.user, payment_method.type_method, form.cleaned_data['description'], amount)

            if not mov:
                messages.error(request, 'Antes de realizar un pago debe Abrir una Caja.')
                return redirect('cashregister_app:cashregister_view')

            Payment.objects.create(
                customer = sale.customer,
                user_made = request.user,
                amount = amount,
                payment_method = payment_method,
                description = f"Pago de duada de venta Nro: {sale.pk}",
                sale = sale,
                movement = mov
            )
            sale.missing_balance -= amount
        if sale.missing_balance <= 0:
            sale.state = "COMPLETADO"
            up_objetives(sale.commision_user, sale)
        sale.save()
        
        messages.success(request, 'Pago realizado con éxito.')
        return redirect('clients_app:customer_detail', pk=service.customer.pk)
    else:
        messages.error(request, "Se produjo un error al realizar pago.")
        return redirect('clients_app:customer_detail', pk=service.customer.pk)
    
        
    
def service_order_entrega(request, pk):
    if request.method == "POST":
        service = ServiceOrder.objects.get(pk=pk)
        service.is_done = True
        service.save()
        messages.success(request, "La Orden de Servicio fue marcada como entregada.")
        return redirect('clients_app:customer_detail', pk=service.customer.pk)

####################### EXPORT CLIENTS LIST #########################

def export_customer_list_to_excel(request):
    # Para la generacion de excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    
    branch_actualy = set_branch_session(request)
    
    list_customer = Customer.objects.get_customers_branch(branch_actualy)

    if not list_customer:
        raise ValueError('No hay clientes para exportar.') # modificar error
    
    # Crear un libro de trabajo de Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Definir estilos personalizados para los encabezados
    header_style = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0b1727', end_color='0b1727', fill_type='solid')

    # Definir los encabezados de las columnas
    headers = ['Fecha de registro', 'Por', 'Nombre', 'Apellido', 'Teléfono', 'DNI', 'Cuenta Corriente']

    # Aplicar estilos a los encabezados y escribir los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_style
        cell.fill = header_fill

    # Modificar el ancho de la columna (ajustar segÃºn tus necesidades)
    #################################################
    try: 
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter
    #################################################
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 25

    # Agregar los datos de los empleados a la hoja de cÃ¡lculo
    for row_num, customer in enumerate(list_customer, 2):
        worksheet.cell(row=row_num, column=1, value=str(customer.created_at.date()))
        worksheet.cell(row=row_num, column=2, value=str(customer.user_made))
        worksheet.cell(row=row_num, column=3, value=customer.first_name)
        worksheet.cell(row=row_num, column=4, value=customer.last_name)
        worksheet.cell(row=row_num, column=5, value=((str(customer.phone_code) + str(customer.phone_number)) if customer.phone_number else "Sin Teléfono"))
        worksheet.cell(row=row_num, column=6, value=customer.dni)
        worksheet.cell(row=row_num, column=7, value= ("Cuanta abierta" if customer.has_credit_account else "Sin cuenta"))

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lista de clientes - Sucursal %s.xlsx' %(branch_actualy.name)

    workbook.save(response)

    return response



########################### RUTINAS PARA PETICIONES AJAX ####################################

def ajax_search_customers(request):
    from django.db.models import Q
    
    branch_actualy = set_branch_session(request)
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        # Obtener el valor de search_term de la solicitud
        search_term = request.GET.get('search_term', '')
        search_terms = search_term.split()
        if not search_term:
            # En caso de que search_term esté vacío, muestra la cantidad de clientes por defecto
            paginate_by = CustomerListView().paginate_by
            customers = Customer.objects.get_customers_branch(branch_actualy).filter(deleted_at=None)[:paginate_by]
        else:
            # Usando reduce para combinar múltiples condiciones con operador OR (|)
            query_conditions = Q()
            for term in search_terms:
                query_conditions |= (
                    Q(first_name__icontains=term) |
                    Q(last_name__icontains=term) |
                    Q(phone_number__icontains=term) |
                    Q(phone_code__icontains=term) |
                    Q(email__icontains=term) |
                    Q(dni__icontains=term)
                )
            customers = Customer.objects.get_customers_branch(branch_actualy).filter(query_conditions)[:50]

        # Crear una lista de diccionarios con los datos de los empleados
        # locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        data = [{
            'id': customer.id,
            'first_name': customer.first_name,
            'last_name': customer.last_name,
            'phone_number': customer.phone_number,
            'phone_code': customer.phone_code,
            'dni': customer.dni,
            'user_made': str(customer.user_made),
            'has_credit_account': 1 if customer.has_credit_account else 0,
            'credit_balance': customer.credit_balance,
            'is_staff': 1 if request.user.is_staff else 0
        } for customer in customers]
        return JsonResponse({'data': data})
    
    return redirect(reverse_lazy('clients_app:customer_view'))


def print_service_order(request, pk): # pk de la orden
    service = ServiceOrder.objects.get(pk=pk)
    branch = service.sale.branch
    # Lógica para obtener el HTML que deseas mostrar en la nueva pestaña
    html_content = "<html><body><h1>Contenido HTML de ejemplo</h1></body></html>"

    # Convertir la cadena en un objeto de fecha
    # locale.setlocale(locale.LC_TIME, 'es_ES.utf8')
    
    format = "%A, %d de %B de %Y"
    
    created_at = service.created_at
    sale_date = created_at
    
    context = {
        'service_order': service,
        'sale_date': sale_date,
        'branch_name': branch.name,
        'branch_phone': branch.phone,
        'branch_address': branch.address,
        'customer': service.customer
    }

    # Genera el HTML en lugar de renderizarlo
    return render(request, 'clients/service_order_print.html', context)