from typing import Any
from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    DetailView, UpdateView, FormView, ListView, DeleteView
)
from django.db import transaction

from django.contrib.auth.mixins import LoginRequiredMixin

from applications.core.mixins import CustomUserPassesTestMixin
from applications.users.models import User
from applications.users.forms import *
from applications.branches.models import Branch_Objetives

from .forms import EmployeeCreateForm, EmployeeUpdateForm
from .models import Employee, Employee_Objetives
from .utils import obtener_nombres_de_campos
from applications.branches.utils import set_branch_session
from django.db.models import Q

# Para la generacion de excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Create your views here.
class EmployeeCreateView(CustomUserPassesTestMixin, FormView): # CREACION DE EMPLEADOS
    template_name = "users/signup.html"
    form_class = EmployeeCreateForm
    success_url = reverse_lazy('employees_app:list_employee')    

    @transaction.atomic
    def form_valid(self, form):
        user = self.request.user
        form.cleaned_data.pop('password2')
        
        
        branch_actualy = set_branch_session(self.request)

        Employee.objects.create(
            user_made = user,
            employment_date = form.cleaned_data.pop('employment_date'),
            jornada = form.cleaned_data.pop('jornada'),
            user = User.objects.create_user(**form.cleaned_data, branch=branch_actualy) # Funcion que crea EMPLEADOS
            )
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al cargar los datos. Por favor, revise los campos.')
        return super().form_invalid(form)


class EmployeeUpdateView(UpdateView):
    model = Employee
    template_name = 'employes/components/employee_update_form.html'
    form_class = EmployeeUpdateForm
    success_url = reverse_lazy('employees_app:list_employee')
    
    def form_valid(self, form):
        employee = form.instance

        if form.is_valid():
            employee.user.first_name = form.cleaned_data.get('first_name')
            employee.user.last_name = form.cleaned_data.get('last_name')
            employee.user.dni = form.cleaned_data.get('id')
            employee.user.address = form.cleaned_data.get('address')
            employee.user.phone_number = form.cleaned_data.get('phone_number')
            employee.user.phone_code = form.cleaned_data.get('phone_code')
            employee.user.birth_date = form.cleaned_data.get('birth_date')

            employee.user.save()

            messages.success(self.request, 'Se actualizo los datos de %s, %s con exito.''.' % (employee.user.last_name, employee.user.first_name))
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al cargar los datos. Por favor, revise los campos.')
        return super().form_invalid(form)


############## UNICA VIEW DISPONIBLE PARA EL USO #############
# Perfil de empleado
class EmployeeProfileView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = 'employes/profile.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_actual = self.request.user
        employee_pk = self.kwargs['pk']
        employee = self.get_object()

        context['is_self'] = True # SI ES ADMIN O EL PROPIO EMPLEADO VIENDO SU PERFIL
        context['objetives'] = Employee_Objetives.objects.filter(employee_id=employee_pk).order_by('created_at')
        context['objetives_branch'] = Branch_Objetives.objects.filter(branch=employee.user.branch).order_by('created_at')

        print(context['objectives'], context['objetives_branch'])

        if not user_actual.is_staff and user_actual.employee_type != self.get_object(): # SI ES UN EMPLEADO QUE ESTA VIENDO OTRO PERFIL
            context['is_self'] = False
        return context

    def get_object(self, queryset=None): 
        """ Obtén el valor del parámetro 'pk' de la URL, este 
        parametro, puede ser la pk de un user, comprobar que esta pk esta relacionada 
        con alguna pk de la tabla users_employee"""
        try:
            employee = Employee.objects.get(pk=self.kwargs['pk'])
        except Employee.DoesNotExist:
            employee = None
        return employee
    

################################## LISTING  ##################################
class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'employes/employee_list_page.html'
    context_object_name = 'employees'
    paginate_by = 25

    def get_queryset(self):
        
        branch_actualy = set_branch_session(self.request)
        return Employee.objects.filter(deleted_at=None, user__branch=branch_actualy)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        exclude_fields_user = [
            "role",
            "phone_code",
            "dni",
            "birth_date",
            "address",
            "id",
            "deleted_at",
            "created_at", 
            "updated_at",
            "date_joined",
            "is_active",
            "is_staff",
            "password",
            "last_login",
            "username",
            "is_superuser",
            "branch",
            "imagen",
        ]
        context['table_column'] = obtener_nombres_de_campos(User, *exclude_fields_user)
        return context
    

########################### DELETE ####################################

class EmployeeDeleteView(LoginRequiredMixin, DeleteView):
    model = Employee
    template_name = 'employes/employee_delete_page.html'
    success_url = reverse_lazy('employees_app:list_employee')
    
    def form_valid(self,form):
        employee = self.get_object()
        user = employee.user
        user.is_active = False  # Realiza la eliminación suave del usuario y por consecuencia, el empleado
        user.delete()
        employee.delete()
        return HttpResponseRedirect(self.get_success_url())

########################### GENERACION EXCEL ####################################

def export_employee_list_to_excel(request):
    branch = request.user.branch

    
    branch_actualy = set_branch_session(request)
    
    queryset = Employee.objects.filter(deleted_at=None, user__branch=branch_actualy)

    # Crear un libro de trabajo de Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Definir estilos personalizados para los encabezados
    header_style = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0b1727', end_color='0b1727', fill_type='solid')

    # Definir los encabezados de las columnas
    exclude_fields_user = ["deleted_at", "created_at", "updated_at"]
    headers = [campo[1] for campo in obtener_nombres_de_campos(Employee, *exclude_fields_user)]

    # Aplicar estilos a los encabezados y escribir los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_style
        cell.fill = header_fill

    # Modificar el ancho de la columna (ajustar según tus necesidades)
    #################################################
    try: 
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter
    #################################################
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 25

    # Agregar los datos de los empleados a la hoja de cálculo
    for row_num, employee in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1, value=str(employee.id))
        worksheet.cell(row=row_num, column=2, value=str(employee.user_made))
        worksheet.cell(row=row_num, column=3, value=employee.user.get_full_name())
        worksheet.cell(row=row_num, column=4, value=employee.employment_date)

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lista de empelados - Sucursal %s.xlsx' %(branch.name)

    workbook.save(response)

    return response


########################### RUTINAS PARA PETICIONES AJAX ####################################

def ajax_search_employee(request):
    branch = request.user.branch

    
    branch_actualy = set_branch_session(request)

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':

        # Obtener el valor de search_term de la solicitud
        search_term = request.GET.get('search_term', '')

        all_employees = Employee.objects.filter(deleted_at=None, user__branch=branch_actualy)

        if not search_term:
            # En caso de que search_term esté vacío, muestra la cantidad de empleados por defecto
            paginate_by = EmployeeListView().paginate_by
            employees = all_employees[:paginate_by]
        else:
            # Usando Q por todos los campos existentes en la tabla first_name, last_name, phone_number, phone_code, email
            employees = all_employees.filter(
                Q(user__first_name__icontains=search_term) |
                Q(user__last_name__icontains=search_term) |
                Q(user__phone_number__icontains=search_term) |
                Q(user__phone_code__icontains=search_term) |
                Q(user__email__icontains=search_term)
            )

        # Crear una lista de diccionarios con los datos de los empleados
        data = [{
            'id': employee.id,
            'first_name': employee.user.first_name,
            'last_name': employee.user.last_name,
            'phone_number': employee.user.phone_number,
            'phone_code': employee.user.phone_code,
            'email': employee.user.email,
            'image_url': employee.user.imagen.url,
            'is_staff': 1 if request.user.is_staff else 0 
        } for employee in employees]
        return JsonResponse({'data': data})
