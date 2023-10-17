from django.db import transaction
from django.db.models import Q
from django.http import HttpResponseRedirect, JsonResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy, reverse
from django.views.decorators.http import require_GET
from django.views.generic import (
    FormView,
    UpdateView,
    DetailView,
    ListView,
    DeleteView
)
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from applications.core.mixins import CustomUserPassesTestMixin
from .models import Brand, Category, Product
from .forms import *
from .utils import *
import math

# Create your views here.

class CategoryCreateView(CustomUserPassesTestMixin, FormView):
    """
    Crear una catogoria nueva para el producto
    """
    form_class = CategoryForm
    template_name = 'products/category_create_page.html'
    success_url = reverse_lazy('products_app:category_list')

    def form_valid(self, form):
        category = form.save(commit=False)
        category.user_made = self.request.user
        category.save()

        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # Para saber si es una peticion AJAX
            new_category_data = {
                'id': category.id,
                'name': category.name
            }
            # Si es una solicitud AJAX, devuelve una respuesta JSON
            return JsonResponse({'status': 'success', 'new_category': new_category_data})
        else:
            # Si no es una solicitud AJAX, llama al método form_valid del padre para el comportamiento predeterminado
            return super().form_valid(form)

class BrandCreateView(LoginRequiredMixin, FormView):
    """
    Crear una marca nueva para los productos
    """
    form_class = BrandForm
    template_name = 'products/brand_create_page.html'
    success_url = reverse_lazy('products_app:brand_list')

    def form_valid(self, form):
        brand = form.save(commit=False)
        brand.user_made = self.request.user
        brand.save()

        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # Para saber si es una peticion AJAX
            new_brand_data = {
                'id': brand.id,
                'name': brand.name
            }
            # Si es una solicitud AJAX, devuelve una respuesta JSON
            return JsonResponse({'status': 'success', 'new_brand': new_brand_data})
        else:
            # Si no es una solicitud AJAX, llama al método form_valid del padre para el comportamiento predeterminado
            return super().form_valid(form)

class FeatureCreateView(CustomUserPassesTestMixin, FormView):
    """
    Crear una característica nueva para el producto
    Previa carga del Tipo de Característica
    """
    form_class = FeatureForm
    template_name = 'products/feature_create_page.html'
    success_url = reverse_lazy('core_app:home')

    def get_context_data(self, **kwargs) :
        context =  super().get_context_data(**kwargs)
        context['type_form'] = FeatureTypeForm()
        return context

    def form_valid(self, form):
        feature = form.save(commit=False)
        feature.user_made = self.request.user
        feature.save()

        for product in form.cleaned_data['products']:
            Product_feature.objects.create(feature=feature, product=product, user_made=self.request.user)

        return super().form_valid(form)
    

class FeatureFullCreateView(FormView):
    """
    Crear una característica nueva para el producto.
        Icluye el TIPO y CARACTERISTICA.
        ¡¡SOLO PARA PETICIONES FETCH!!
    """
    form_class = FeatureForm_toWizard

    def form_valid(self, form):
        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # Para saber si es una peticion AJAX
            feature, created_ft = validate_exists_feature_full(form, self.request.user)
            type = form.cleaned_data['type']
            
            if not created_ft: # Pregunta si ya existe el tipo solo
                return JsonResponse({'status': 'error', 'error_message': f'Por favor cree la nueva característica desde la columna "{type}".'})

            new_feature_data = {
                'id':  feature.id,
                'name': feature.__str__(),
            }
            # Si es una solicitud AJAX, devuelve una respuesta JSON
            return JsonResponse({'status': 'success', 'new_feature': new_feature_data})

        
class FeatureUnitCreateView(FormView):
    """
    Añade una caracteristica a un TIPO seleccionado.
        ¡¡SOLO PARA PETICIONES FETCH!!
    """
    form_class = FeatureForm_toWizard

    def form_valid(self, form):
        if len(form.cleaned_data['value']) < 1:
            return JsonResponse({'status': 'error', 'error_message': 'Debe ingresar por lo menos 1 caracter.'})

        type = Feature_type.objects.get(name=form.cleaned_data['type'])

        feature, created = Feature.objects.get_or_create(
            value = form.cleaned_data['value'],
            type = type,
            user_made = self.request.user
        )
        if not created: 
            return JsonResponse({'status': 'error', 'error_message': 'Los datos ingresados ya existen.'})
        
        new_feature_data = {
                'id':  feature.id,
                'name': feature.value,
                'type': feature.type.name,
            }
            # Si es una solicitud AJAX, devuelve una respuesta JSON
        return JsonResponse({'status': 'success', 'new_feature_unit': new_feature_data})


class FeatureTypeCreateView(CustomUserPassesTestMixin, FormView): # TIPO DE CARACTERISTICA (1)
    """
    Crear un tipo para la caracteristica nueva para el producto
        Es lo primero que se crea, luego se crea la caracteristica (tipo -> caracteristica)
    """
    form_class = FeatureTypeForm
    template_name = 'products/featureType_create_page.html'
    success_url = reverse_lazy('products_app:new_feature')

    def form_valid(self, form):
        feature_type = form.save(commit=False)
        feature_type.user_made = self.request.user
        feature_type.save()
                
        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # Para saber si es una peticion AJAX
            new_type_data = {
                'id': feature_type.id,
                'name': feature_type.name
            }
            # Si es una solicitud AJAX, devuelve una respuesta JSON
            return JsonResponse({'status': 'success', 'new_type': new_type_data})
        else:
            # Si no es una solicitud AJAX, llama al método form_valid del padre para el comportamiento predeterminado
            return super().form_valid(form)


#################### FORMULARIO WIZARD #####################
class ProductCreateView(CustomUserPassesTestMixin, FormView):
    """
    Vista para crear un producto completo, con caracteristicas.
    """
    form_class = ProductForm
    template_name = 'products/wizard_create_product.html'
    success_url = reverse_lazy('products_app:product_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_form'] = CategoryForm
        context['brand_form'] = BrandForm
        context['feature_form'] = FeatureForm_toWizard
        return context
    
    def get_form_kwargs(self):
        user = self.request.user
        branch_actualy = self.request.session.get('branch_actualy')
        if user.is_staff and branch_actualy:
            branch_actualy = Branch.objects.get(id=branch_actualy)
            branch = branch_actualy
        else:
            branch = user.branch

        kwargs = super().get_form_kwargs()
        kwargs['branch'] = branch
        return kwargs

    @transaction.atomic
    def form_valid(self, form):
        user = self.request.user

        if form.is_valid():
            if user.is_staff:
                branch_actualy = self.request.session.get('branch_actualy')
                branch_actualy = Branch.objects.get(id=branch_actualy)
                branch = branch_actualy
            else:
                branch = user.branch
            
            percentage = 1.26
            multiplicador = 3
            packaging = 580

            product = form.save(commit=False)
            suggested_price = Decimal((float(product.cost_price) * percentage) * multiplicador)

            if product.has_eyeglass_frames:
                suggested_price = suggested_price + Decimal(packaging)

            sale_price = math.ceil(suggested_price / 50) * 50
            product.sale_price = sale_price

            product.user_made = self.request.user
            product.branch = branch
            product.save()
            form_in_out_features(form, product, self.request.user)

        feature_formset = FeatureFormSet(self.request.POST, prefix='variants')

        if feature_formset.is_valid():
            form_create_features_formset(self.request.user, product, feature_formset)
        
        source_route = self.request.GET.get('source_route', '')  # Obtén el valor del parámetro 'source_route' de la URL
        supplier_pk = self.request.GET.get('supplier_pk', '')  # Obtén el valor del parámetro 'supplier_pk' de la URL

        # Determina la URL de redirección según la ruta de origen
        if source_route == 'supplier_create':
            redirect_url = reverse('suppliers_app:new_supplier')
        elif source_route == 'supplier_update':
            redirect_url = reverse('suppliers_app:update_supplier', args=[supplier_pk])
        else:
            redirect_url = reverse('products_app:product_list')  # Ruta predeterminada si no se especifica la ruta de origen

        return HttpResponseRedirect(redirect_url)

    def form_invalid(self, form):
        messages.error(self.request, 'Hubo un error al cargar los datos. Por favor, revise los campos.')
        return super().form_invalid(form)
    
    
####################### UPDATES #####################

class ProductUpdateView(CustomUserPassesTestMixin, UpdateView):
    """
    Vista para Actualizar un producto.
    """
    model = Product
    form_class = ProductForm
    template_name = 'products/wizard_create_product.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category_form'] = CategoryForm(instance=self.object.category)
        context['brand_form'] = BrandForm(instance=self.object.brand)
        context['feature_form'] = FeatureForm_toWizard
        context['update'] = 1
        return context

    @transaction.atomic
    def form_valid(self, form):
        # Si es un nuevo producto, simplemente guarda el formulario
        product = form.instance
        product.user_made = self.request.user

        feature_formset = FeatureFormSet(self.request.POST, prefix='variants')
        form_in_out_features(form, product, self.request.user)
        if feature_formset.is_valid():
            form_create_features_formset(self.request.user, product, feature_formset)
        self.success_url = reverse_lazy('products_app:product_detail', kwargs={'pk': self.get_object().pk})
        return super().form_valid(form)


class CategoryUpdateView(CustomUserPassesTestMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'products/category_update_page.html'
    success_url = reverse_lazy('products_app:category_list')
    
    def form_valid(self, form):
        category = form.save(commit=False)
        category.user_made = self.request.user
        category.save()
                
        return super().form_valid(form)


class BrandUpdateView(CustomUserPassesTestMixin, UpdateView):
    model = Brand
    form_class = BrandForm
    template_name = 'products/brand_update_page.html'
    success_url = reverse_lazy('products_app:brand_list')

    def form_valid(self, form):
        form.instance.user_made = self.request.user
        return super().form_valid(form)

class FeatureUpdateView(CustomUserPassesTestMixin, UpdateView):
    model = Feature
    form_class = FeatureForm
    template_name = 'products/feature_update_page.html'
    success_url = reverse_lazy('products_app:feature_list')

    def form_valid(self, form):
        form.instance.user_made = self.request.user
        return super().form_valid(form)

class FeatureTypeUpdateView(CustomUserPassesTestMixin, UpdateView):
    model = Feature_type
    form_class = FeatureForm
    template_name = 'products/featureType_update_page.html'
    success_url = reverse_lazy('core_app:homefeature_type_list')

    def form_valid(self, form):
        form.instance.user_made = self.request.user
        return super().form_valid(form)

################## LIST ######################

class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'products/product_list_page.html'
    context_object_name = 'products'
    paginate_by = 25

    def get_queryset(self):
        user = self.request.user
        if user.is_staff:
            branch_actualy = self.request.session.get('branch_actualy')
            branch_actualy = Branch.objects.get(id=branch_actualy)
            branch = branch_actualy
        else:
            branch = user.branch

        return Product.objects.filter(branch=branch, deleted_at=None)

    def get_context_data(self, **kwargs):
        pk=self.request.user.pk
        context = super().get_context_data(**kwargs)
        
        exclude_fields = ["id", "deleted_at", "created_at", "updated_at","cost_price","suggested_price", "user_made", "branch", "has_eyeglass_frames", "promotion"]
        context['table_column'] = obtener_nombres_de_campos(Product, *exclude_fields)
        context['features']=Product_feature.objects.filter(product_id=pk)
        return context


class BrandListView(LoginRequiredMixin, ListView):
    model = Brand
    template_name = 'products/brand_list_page.html'
    context_object_name = 'brands'
    
    def get_queryset(self):
        # Filtra los productos que no han sido eliminados suavemente
        return Brand.objects.filter(deleted_at=None)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['table_column'] = obtener_nombres_de_campos(Brand,"id","deleted_at", "created_at", "updated_at")
        return context


class CategoryListView(LoginRequiredMixin, ListView):
    model = Category
    template_name = 'products/category_list_page.html'
    context_object_name = 'categories'

    def get_queryset(self):
        # Filtra los productos que no han sido eliminados suavemente
        return Category.objects.filter(deleted_at=None)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['table_column'] = obtener_nombres_de_campos(Category,"id","deleted_at", "created_at", "updated_at") 
        return context


class FeatureTypeListView(LoginRequiredMixin, ListView):
    model = Feature_type
    template_name = 'products/feature_type_page.html'
    context_object_name = 'feature_type'

    def get_queryset(self):
        # Filtra los productos que no han sido eliminados suavemente
        return Feature_type.objects.filter(deleted_at=None)

class FeatureListView(LoginRequiredMixin, ListView):
    model = Feature
    template_name = 'products/feature_page.html'
    context_object_name = 'feature'

    def get_queryset(self):
        # Filtra los productos que no han sido eliminados suavemente
        return Feature.objects.filter(deleted_at=None)

#################### DETAILS #####################

class ProductDetailView(LoginRequiredMixin, DetailView):
    model = Product
    template_name = 'products/product_page.html'
            
    def get(self, request, *args, **kwargs):
        user = request.user
        branch = user.branch
        object = self.get_object()

        if not user.is_staff and not object.branch == branch: 
            # El usuario no tiene permiso para ver este producto
            messages.error(request, 'Lo sentimos, no puedes ver este producto.')
            return redirect('products_app:product_list')  # Reemplaza 'nombre_de_tu_vista_product_list' por el nombre correcto de la vista de lista de productos
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        features_list = Product.objects.get_feature_types_and_values(self.get_object())
        context['features'] = features_list
        return context
    
    
class CategoryDetailView(LoginRequiredMixin, DetailView):
    model = Category
    template_name = 'products/category_page.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['products'] = Product.objects.filter(category=self.object, deleted_at=None)
        # AGREGAR MAS DETALLES RELACIONADOS
        return context


class BrandDetailView(LoginRequiredMixin, DetailView):
    model = Brand
    template_name = 'products/brand_page.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['products'] = Product.objects.filter(brand=self.object, deleted_at=None)
        # AGREGAR MAS DETALLES RELACIONADOS
        return context
    


########################### DELETE ####################################
 # FALTO COMPLETAR LA LOGICA PARA EL BORRADO SUEAVE DE LAS INSTANCIAS REALCIONADAS
    # QUE PASA SI BORRO UNA FEATURE QUE YA ESTA RELACIONADA PREVIAMENTE CON ALGUN PRODUCTO ? (PROTECTED NO DEJARA QUE BORRE)
class CategoryDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = Category
    template_name = 'products/category_delete_page.html'
    success_url = reverse_lazy('products_app:category_list')
    

class BrandDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = Brand
    template_name = 'products/brand_delete_page.html'
    success_url = reverse_lazy('products_app:brand_list')
    

class ProductDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = Product
    template_name = 'products/product_delete_page.html'
    success_url = reverse_lazy('products_app:product_list')


class FeatureDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = Feature
    template_name = 'products/category_form.html'
    success_url = reverse_lazy('core_app:home')
    

class FeatureTypeDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = Feature_type
    template_name = 'products/category_form.html'
    success_url = reverse_lazy('products_app:new_feature')
    


###################### SEARCH #########################
class ProductSearchView(ListView):
    """
    Buscador de productos para el Punto de Venta (PoS)
        ¡¡SOLO PARA PETICIONES FETCH!!
    """

    def get_queryset(self):
        user = self.request.user
        branch_actualy = self.request.session.get('branch_actualy')
        if user.is_staff and branch_actualy:
            branch_actualy = Branch.objects.get(id=branch_actualy)
            branch = branch_actualy
        else:
            branch = user.branch

        query = self.request.GET.get('q')
        if query:
            # Convertir la consulta a minúsculas
            query = query.lower()

            # Realizar una búsqueda insensible a mayúsculas/minúsculas y que contenga la palabra en los campos 'name' y 'barcode'
            queryset = Product.objects.filter(
                Q(name__icontains=query) |
                Q(name__icontains=query.capitalize()) |
                Q(name__icontains=query.upper()) |
                Q(name__icontains=query.lower()) |
                Q(name__icontains=query.title()) |
                Q(barcode__icontains=query),
                branch=branch)
        else:
            queryset = Product.objects.all()

        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        data = [{
            'id': product.pk, 
            'name': product.name, 
            'description': product.description, 
            'price': product.sale_price, 
            'category': product.category.name, 
            'brand': product.brand.name,
            'barcode': product.barcode 
            } for product in queryset]

        return JsonResponse({'products': data})


#################### Actualizar Precio ###############

class UpdatePriceView(CustomUserPassesTestMixin, FormView):
    template_name = 'products/update_price_advanced.html'
    form_class = UpdatePriceForm
    success_url = reverse_lazy('products_app:product_list')

    def form_valid(self, form, *args, **kwargs):
        if form.is_valid():
            #me devuelve el objeto de categoria
            brands_up = form.cleaned_data['brand']
            categories_up = form.cleaned_data['category']
            percentage = form.cleaned_data['percentage']

            # Crear conjuntos para las marcas y las categorías
            product_brand = set()
            product_category = set()

            for brand in brands_up:
                product_brand.update(brand.product_brand.all())

            for category in categories_up:
                product_category.update(category.product_category.all())

            # Ahora, product_brand y product_category contienen objetos únicos sin duplicados
            products = product_category | product_brand

            for product in products:
                new_price = product.sale_price * (1 + (percentage / 100))
                new_price = math.ceil(new_price / 50) * 50
                product.sale_price = Decimal(new_price)
                product.save()

        return super().form_valid(form)

    def form_invalid(self, form):
        print('\n\n\n\n', form.errors)
        return super().form_invalid(form)
    

@require_GET
def search_categories_or_brands(request):
    search_term = request.GET.get('search_term', '')

    results = []

    categories = Category.objects.filter(name__icontains=search_term)[:5]
    brands = Brand.objects.filter(name__icontains=search_term)[:5]

    combined_results = list(categories) + list(brands)
    results = [{'id': item.id, 'name': item.name, 'form_name': 'brand' if item.__class__ is Brand else 'category'} for item in combined_results]

    return JsonResponse(results, safe=False)


###################### EXPORT PRODUCTS LIST ########################

def export_products_list_to_excel(request):
    # Para la generacion de excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    branch_actualy = request.session.get('branch_actualy')
    if request.user.is_staff and branch_actualy:
        branch_actualy = Branch.objects.get(id=branch_actualy)
        branch = branch_actualy
    else:
        branch = request.user.branch
    
    list_products = Product.objects.get_products_branch(branch)

    if not list_products:
        raise ValueError('No hay productos para exportar.') # modificar error
    
    # Crear un libro de trabajo de Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Definir estilos personalizados para los encabezados
    header_style = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0b1727', end_color='0b1727', fill_type='solid')

    # Definir los encabezados de las columnas
    exclude_fields_user = ["deleted_at", "updated_at", "id", "branch", "user_made", "has_eyeglass_frames", "promotion"]
    headers = [campo[1] for campo in obtener_nombres_de_campos(Product, *exclude_fields_user)]

    if 'created at' in headers:
        index = headers.index('created at')
        headers[index] = "Fecha de registro"
    # Aplicar estilos a los encabezados y escribir los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_style
        cell.fill = header_fill

    # Modificar el ancho de la columna (ajustar segÃºn tus necesidades)
    #################################################
    try: 
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter
    #################################################
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 25

    # Agregar los datos de los empleados a la hoja de cÃ¡lculo
    for row_num, product in enumerate(list_products, 2):
        worksheet.cell(row=row_num, column=1, value=str(product.created_at.date()))
        worksheet.cell(row=row_num, column=2, value=product.name)
        worksheet.cell(row=row_num, column=3, value=str(product.barcode))
        worksheet.cell(row=row_num, column=4, value=str(product.cost_price))
        worksheet.cell(row=row_num, column=5, value=str(product.sale_price))
        worksheet.cell(row=row_num, column=6, value=product.description)
        worksheet.cell(row=row_num, column=7, value=str(product.stock))
        worksheet.cell(row=row_num, column=8, value=product.category.name)
        worksheet.cell(row=row_num, column=9, value=product.brand.name)

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lista de productos - Sucursal %s.xlsx' %(branch.name)

    workbook.save(response)

    return response


################ SEARCH PRODUCTS AJAX ################

def ajax_search_products(request):
    branch = request.user.branch

    branch_actualy = request.session.get('branch_actualy')
    if request.user.is_staff and branch_actualy:
        branch = Branch.objects.get(id=branch_actualy)

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':

        # Obtener el valor de search_term de la solicitud
        search_term = request.GET.get('search_term', '')

        if not search_term:
            # En caso de que search_term esté vacío, muestra la cantidad de empleados por defecto
            paginate_by = ProductListView().paginate_by
            products = Product.objects.get_products_branch(branch)[:paginate_by]
        else:
            # Usando Q por todos los campos existentes en la tabla
            products = Product.objects.get_products_branch(branch).filter(
                Q(name__icontains=search_term) |
                Q(barcode__icontains=search_term) |
                Q(description__icontains=search_term) |
                Q(category__name__icontains=search_term) |
                Q(brand__name__icontains=search_term)
            )[:5]
        # Crear una lista de diccionarios con los datos de los empleados
        data = [{
            'id': product.id,
            'name': product.name,
            'barcode': product.barcode,
            'sale_price': product.sale_price,
            'description': product.description,
            'stock': product.stock,
            'category': product.category.name,
            'brand': product.brand.name,
            'is_staff': 1 if request.user.is_staff else 0
        } for product in products]
        return JsonResponse({'data': data})