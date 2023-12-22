import copy
import locale
from typing import Any
from django.db.models import Q
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import redirect, render
from django.views.generic import *
from django.template import loader
from django.db import transaction
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin

#Importaciones de la app
from branches.utils import set_branch_session
from clients.forms import *
from promotions.models import Promotion
from cashregister.utils import create_in_movement, obtener_nombres_de_campos
from core.mixins import CustomUserPassesTestMixin
from cashregister.models import CashRegister, Movement

from .utils import *
from .models import *
from .forms import *
from django.views.generic import ListView

# Create your views here.

class PointOfSaleView(LoginRequiredMixin, FormView):
    form_class = OrderDetailFormset
    template_name = 'sales/point_of_sale_page.html'

    def get(self, request: HttpRequest, *args: str, **kwargs: Any) -> HttpResponse:
        branch_actualy = set_branch_session(self.request)
        
        cashregister = CashRegister.objects.filter(is_close=False, branch=branch_actualy).last()
        if not cashregister:
            messages.error(self.request, 'Antes de realizar una Venta, debe Abrir una Caja.')
            return redirect('cashregister_app:cashregister_view')
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        branch_actualy = set_branch_session(self.request)

        context['sale_form'] = SaleForm(branch=branch_actualy)
        context['payment_method_formset'] = PaymentMethodsFormset
        context['branch_selected'] = branch_actualy.name
        context['customer_form'] = CustomerForm
        context['payment_method_form'] = PaymentMethodForm
        
        return context

    @transaction.atomic
    def form_valid(self, form):
        branch_actualy = set_branch_session(self.request)

        try: 
            CashRegister.objects.filter(
                is_close = False,
                branch = branch_actualy,
            ).last()
        except CashRegister.DoesNotExist:
            messages.error(self.request, 'Antes de realizar una Venta, debe Abrir una Caja.')
            return redirect('cashregister_app:cashregister_view')

        formsets = form
        saleform = SaleForm(data=self.request.POST, branch=branch_actualy)
        payment_methods = PaymentMethodsFormset(self.request.POST)
                
        order_details = []
        all_products_to_sale = []
        real_price_promo = []
        wo_promo = []
        subtotal = 0

        if saleform.is_valid():
            sale = saleform.save(commit=False)
            sale.branch = branch_actualy

            if not saleform.cleaned_data['commision_user']:
                messages.error(self.request, "Debe seleccionar un comisionista.")
                return super().form_invalid(form)
            customer = saleform.cleaned_data['customer']

            cuenta_corriente_presente = any(
                'Cuenta Corriente' in form.cleaned_data.get('name', '') for form in payment_methods
                if form.is_valid()
            )

            if customer and not customer.has_credit_account and cuenta_corriente_presente:
                messages.error(self.request, "El metodo de pago 'Cuenta Corriente' no esta habilitado para el cliente seleccionado.")
                return super().form_invalid(form)
                
            if not customer and cuenta_corriente_presente:
                messages.error(self.request, "El metodo de pago 'Cuenta Corriente' no esta habilitado para el cliente seleccionado.")
                return super().form_invalid(form)
            

            amount = saleform.cleaned_data.pop('amount')
            discount_sale = saleform.cleaned_data['discount']
            surcharge_sale = saleform.cleaned_data['surcharge']
        
        else:
            print(saleform.errors)
            messages.error(self.request, next(iter(saleform.errors.values()))[0])
            return super().form_invalid(form) 
            
        promotions_active = Promotion.objects.filter(is_active=True, branch=branch_actualy, deleted_at=None)
        promotional_products = {promotion: [(promotion.discount)] for promotion in promotions_active}

        # Procesa los datos del formset
        for formset in formsets:
            if not formset.cleaned_data['product']:
                messages.error(self.request, "No se ha seleccionado ningun producto.")
                return super().form_invalid(form)
            
            if formset.is_valid():
                product = formset.cleaned_data['product']
                quantity = formset.cleaned_data['quantity']
                all_products_to_sale.append(product)

                # Asegurarse de que la cantidad que se va a vender sea menor o igual al stock actual
                if product.stock < quantity:
                    messages.warning(self.request, f"No hay suficiente stock disponible para {product.name}.")
                    return super().form_invalid(form)

        cristal = find_cristal_product(all_products_to_sale)
        contacto = find_contacto_product(all_products_to_sale)
        armazon = find_armazons_product(all_products_to_sale)
        
        if cristal and armazon:
            if cristal and not customer:
                messages.error(self.request, "Seleccione un Cliente antes de vender un Cristal.")
                return super().form_invalid(form)
            
        elif cristal and not armazon:
            messages.error(self.request, "Seleccione un Armazón antes de vender un Cristal.")
            return super().form_invalid(form)
        
        if contacto and not customer:
            messages.error(self.request, "Seleccione un Cliente antes de vender un Lente de Contacto.")
            return super().form_invalid(form)

        for formset in formsets:
            if formset.is_valid():
                subtotal += get_total_and_products(formset)
                order_details.append(process_formset(formset, promotional_products, wo_promo))


        promotional_products_clone = copy.copy(promotional_products)
        # Ordena los productos en cada promoción por precio
        for promotion, products_with_discountPromo in promotional_products.items():
            process_promotion(promotional_products_clone, promotion, products_with_discountPromo, real_price_promo)
                
        set_amounts_sale(sale, subtotal, wo_promo, real_price_promo, discount_sale, surcharge_sale)

        if (cristal or contacto) and amount < sale.total/2 and not customer.has_credit_account: # Se lleva un cristal o lente de contacto, pero el monto pagado es menor al 50%
            messages.warning(self.request, "El pago debe ser mayor al 50% del total.")
            return super().form_invalid(form)
        
        if not (cristal or contacto) and amount < sale.total and (customer is None or not customer.has_credit_account):
            messages.warning(self.request, "Los pagos parciales solo están habilitados para la Venta con Cristales o Lentes de Contacto.")
            return super().form_invalid(form)

        process_customer(customer, sale, payment_methods, Decimal(sale.total), cristal, contacto, amount, self.request)
        if sale.state == 'COMPLETADO':
            up_objetives(sale.commision_user, sale)

        for formset in formsets:
            if formset.is_valid():
                update_stock(formset)

        for order in order_details:
            order.sale = sale
            order.save()
        
        messages.success(self.request, "Se ha generado la venta con éxito!")
        return HttpResponseRedirect(reverse_lazy('sales_app:sale_detail_view', kwargs={'pk': sale.id}))

    def form_invalid(self, form):
        messages.error(self.request, "Error. Verifique los datos.")
        return super().form_invalid(form)


class PaymentMethodCreateView(CustomUserPassesTestMixin, FormView):
    """
    Crear una catogoria nueva para el producto
    """
    form_class = PaymentMethodForm
    template_name = 'sales/payment_method_create_page.html'
    success_url = reverse_lazy('sales_app:payment_method_view')

    def form_valid(self, form):
        payment_method  = form.save(commit=False)
        payment_method.user_made = self.request.user
        payment_method.save()

        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest': # Para saber si es una peticion AJAX
            new_payment_method_data = {
                'id': payment_method.id,
                'name': payment_method.__str__(),
            }
            # Si es una solicitud AJAX, devuelve una respuesta JSON
            return JsonResponse({'status': 'success', 'new_payment_method': new_payment_method_data})
        else:
            # Si no es una solicitud AJAX, llama al método form_valid del padre para el comportamiento predeterminado
            return super().form_valid(form)
    
    @transaction.atomic
    def form_invalid(self, form):
        if self.request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message':next(iter(form.errors.values()))[0]})
        return super().form_invalid(form)
    
    
class PaymentMethodView(CustomUserPassesTestMixin, ListView):
    """
    Listar todas las categorias de productos
    """
    model = PaymentMethod
    template_name = 'sales/payment_method_list_page.html'
    paginate_by = 50

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['payment_method_form'] = PaymentMethodForm
        return context
    
    
######################## SALES #############################

class SalesListView(LoginRequiredMixin, ListView):
    template_name = 'sales/sale_page.html'
    model = Sale
    paginate_by = 50
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Aquí se recupera la caja de la sucursal correspondiente al usuario logueado
        
        branch_actualy = set_branch_session(self.request)

        sales = Sale.objects.filter(branch=branch_actualy, deleted_at=None).order_by('-created_at')

        context['sales'] = sales
        context['table_column'] = obtener_nombres_de_campos(Sale,
            "id",
            "invoice", 
            "receipt",
            "refund_date",
            "branch",
            "deleted_at", 
            "discount",
            "discount_extra",
            "updated_at",
            "subtotal",
            "created_at"
            )
        
        return context
    

class SaleDetailView(LoginRequiredMixin, DetailView):
    template_name = 'sales/sale_detail_page.html'
    model = Sale
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #detalles de venta en la venta que viene por url
        context['sale'] = sale = Sale.objects.get(id=self.kwargs['pk'])
        context['sale_subtotal'] = sale.subtotal
        context['sale_discount_amount'] = context['sale_subtotal'] * Decimal(sale.discount/100)
        subtotal_discount = context['sale_subtotal'] - context['sale_discount_amount'] 
        context['sale_surcharge'] = subtotal_discount * Decimal(sale.surcharge/100)
        context['sale_total'] = sale.total
        # Ordenes de detalle de la venta ...
        context['sale_details'] = OrderDetail.objects.filter(sale=sale)
        
        context['cristales'] = find_cristal_product(None, sale)
        context['contactos'] = find_contacto_product(None, sale)
        
        armazones = find_armazons_product(None, sale)

        context['order_service'] = {
            'service': ServiceOrderForm(kwargs = armazones),
            'pupilar': InterpupillaryForm,
            'correction': CorrectionForm,
            'material': MaterialForm,
            'color': ColorForm,
            'cristal': CristalForm,
            'tratamiento': TratamientForm,
        }
        print(sale.created_at)
        context['select'] = SelectFacturaFrom

        context['order_serivices'] = sale.service_order.all()

        context['pay_form'] = TypePaymentMethodForm
        return context
    
class SaleDeleteView(CustomUserPassesTestMixin, DeleteView):
    model = Sale
    template_name = 'sales/sales_delete_page.html'
    success_url = reverse_lazy('sales_app:sales_list_view')
    
    def post(self, request, *args, **kwargs):
        sale = self.get_object()
        payments = sale.sale_payment.all()
        for payment in payments:
            try: 
                mov = payment.movement

                cashregister = mov.cash_register
                cashregister.final_balance -= mov.amount
                cashregister.save()

                mov.delete()
            except:
                print('Moviemiento error')
            payment.delete()
        
        order_detaill = sale.order_detaill.all()
        
        for order_detail in order_detaill:
            product = order_detail.product
            product.stock += order_detail.quantity
            product.save()
            order_detail.delete()
            
        sale_date = sale.created_at.date()
        try:
            employee = sale.commision_user
            
            objetive = employee.employee_objetives.filter(
                objetive__start_date__lte=sale_date, 
                objetive__exp_date__gte=sale_date, deleted_at=None)
            if objetive:
                objetive = objetive.first()
                objetive.accumulated -= sale.total
                objetive.save()

            objetive_branch = Branch_Objetives.objects.filter(objetive__branch = sale.user_made.branch,
                objetive__start_date__lte=sale_date, 
                objetive__exp_date__gte=sale_date, deleted_at=None)
            if objetive_branch:
                objetive_branch = objetive_branch.first()
                objetive_branch.accumulated -= sale.total
                objetive_branch.save()
        except Exception as e:
            print('Objetivo error: ', e)
        
        customer = sale.customer
        if customer.has_credit_account:
            customer.credit_balance -= sale.total
            customer.save()
        
        return super().post(request, *args, **kwargs)
            
#------- VISTAS BASADAS EN FUNCIONES PARA PETICIONES AJAX -------#

def show_invoice(request, pk):
    sale = Sale.objects.get(id=pk)
    customer = sale.customer

    payments = Payment.objects.filter(sale=sale)

    if payments.exists():
        # Si hay al menos un pago, toma el primero (puedes ajustar la lógica según tus necesidades)
        payment = payments.first()
    
    order_details = sale.order_detaill.filter(sale=sale)
    order_details_template = []
    subtotal = []

    for order in list(order_details):
        order_details_template.append((order, f'{Decimal(order.price)*Decimal(1-order.discount/100):.2f}'))
        subtotal.append(order.price * order.quantity)

    # Convertir la cadena en un objeto de fecha
    # locale.setlocale(locale.LC_TIME, 'es_ES.utf8')

    format = "%A, %d de %B de %Y"
    
    created_at = sale.created_at
    sale_date_str = created_at.strftime(format)

    disocunt_amount = sale.subtotal * Decimal(sale.discount/100)
    surcharge_amount = (sale.subtotal - disocunt_amount) * Decimal(sale.surcharge/100)

    context = {
        'customer': customer,
        'total': f'{sale.total:.2f}',
        'order_details': order_details_template,
        'subtotal': f'{Decimal(sum(subtotal)):.2f}',
        'seler': sale.commision_user,
        'promo': f'{sale.total:.2f}',
        'discount': f'{sale.discount:.2f}' if sale.discount > 0 else None,
        'discount_amount': f'{disocunt_amount:.2f}' if disocunt_amount > 0 else None,
        'discount_extra': f'{sale.discount_extra:.2f}' if sale.discount_extra > 0 else None,
        'payment_method': payment.payment_method.name,
        'pay': f'{sale.total - sale.missing_balance:.2f}',
        'missing_balance': f'{sale.missing_balance:.2f}',
        'date': created_at,
        'payments': payments,
        'branch': sale.branch,
        'surcharge': f'{sale.surcharge:.2f}' if sale.surcharge > 0 else None,
        'surcharge_amount': f'{surcharge_amount:.2f}' if surcharge_amount > 0 else None,
    }

    return render(request, 'sales/components/comprobante_pago.html', context)


def show_factura(request, pk):
    sale = Sale.objects.get(id=pk)

    customer = sale.customer
    receipt = sale.receipt

    payments = Payment.objects.filter(sale=sale)

    if payments.exists():
        # Si hay al menos un pago, toma el primero (puedes ajustar la lógica según tus necesidades)
        payment = payments.first()
        payment_method = payment.payment_method.name
    else:
        # Manejar el caso en el que no hay pagos
        payment_method = None

    order_details = sale.order_detaill.filter(sale=sale)
    order_details_template = []
    subtotal = []

    for order in list(order_details):
        order_details_template.append((order, f'{Decimal(order.price)*Decimal(1-order.discount/100):.2f}'))
        subtotal.append(order.price * order.quantity)

    # Convertir la cadena en un objeto de fecha
    # # locale.setlocale(locale.LC_TIME, 'es_ES.utf8')

    format = "%A, %d de %B de %Y"
    
    created_at = sale.created_at
    sale_date_str = created_at.strftime(format)

    disocunt_amount = sale.subtotal * Decimal(sale.discount/100)
    surcharge_amount = (sale.subtotal - disocunt_amount) * Decimal(sale.surcharge/100)

    print(surcharge_amount)
    context = {
        'customer': customer,
        'total': f'{sale.total:.2f}',
        'order_details': order_details_template,
        'subtotal': f'{Decimal(sum(subtotal)):.2f}',
        'seler': sale.commision_user,
        'promo': f'{sale.total:.2f}',
        'discount': f'{sale.discount:.2f}' if sale.discount > 0 else None,
        'discount_amount': f'{disocunt_amount:.2f}' if disocunt_amount > 0 else None,
        'discount_extra': f'{sale.discount_extra:.2f}' if sale.discount_extra > 0 else None,
        'payment_method': payment.payment_method.name,
        'pay': f'{sale.total - sale.missing_balance:.2f}',
        'missing_balance': f'{sale.missing_balance:.2f}',
        'time': datetime.now().strftime('%H:%M'),     # Formato: HH:MM
        'receipt': receipt,
        'payments': payments,
        'surcharge': f'{sale.surcharge:.2f}' if sale.surcharge > 0 else None,
        'surcharge_amount': f'{surcharge_amount:.2f}' if surcharge_amount > 0 else None,
        
    }

    return render(request, 'sales/components/factura.html', context)


def gen_factura(request, pk):
    if request.method == "POST":
        sale = Sale.objects.get(id=pk)
        
        select = SelectFacturaFrom(request.POST)
        if select.is_valid():
            option = select.cleaned_data["select"]
            error = switch_invoice_receipt(option, sale, sale.branch.pos_afip)
            if error is not None:
                messages.error(request, error)
                return redirect('sales_app:sale_detail_view', pk=pk)
        else:
            messages.error(request, 'Se produjo un error al generar la factura.')
    return redirect('sales_app:sale_detail_view', pk=pk)

    

################ SEARCH MOVEMENTS AJAX ################

def ajax_search_sales(request):
    branch_actualy = set_branch_session(request)

    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':

        # Obtener el valor de search_term de la solicitud
        search_term = request.GET.get('search_term', '')
        search_terms = search_term.split()
        if not search_term:
            # En caso de que search_term esté vacío, muestra la cantidad de empleados por defecto
            paginate_by = SalesListView().paginate_by
            sales = Sale.objects.all().filter(deleted_at = None)[:paginate_by]
        else:
            # Usando reduce para combinar múltiples condiciones con operador OR (|)
            query_conditions = Q()
            for term in search_terms:
                query_conditions |= (
                    Q(total__icontains=term) |
                    Q(missing_balance__icontains=term) |
                    Q(created_at__icontains=term) |  # Cambié created_at__date__icontains a created_at__icontains
                    Q(state__icontains=term) |
                    Q(user_made__first_name__icontains=term) |
                    Q(user_made__last_name__icontains=term) |
                    Q(customer__first_name__icontains=term) |
                    Q(customer__last_name__icontains=term) |
                    Q(customer__dni__icontains=term)
                )

            sales = Sale.objects.filter(deleted_at=None, branch=branch_actualy).filter(query_conditions)[:40]
        # Crear una lista de diccionarios con los datos de los empleados
        # locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        data = [{
            'id': sale.id,
            'total': sale.total,
            'missing_balance': sale.missing_balance,
            'date_time_sale': sale.created_at.time().strftime('%H:%M'),
            'date_sale': sale.created_at.date().strftime('%Y-%m-%d'),
            'state': sale.state,
            'customer': str(sale.customer),
            'customer_id': sale.customer.id,
            'user_made': str(sale.user_made),
            'is_staff': 1 if request.user.is_staff else 0
        } for sale in sales]
        # locale.setlocale(locale.LC_TIME, '')
        return JsonResponse({'data': data})
    

def set_serviceOrder_onSale(request, pk):
    sale = Sale.objects.get(pk=pk)
    customer = sale.customer

    service_order = process_service_order(request, customer)
    service_order.sale = sale
    service_order.save()

    return HttpResponseRedirect(reverse_lazy('sales_app:sale_detail_view', kwargs={'pk': pk}))


def pay_missing_balance(request, pk):
    if request.method == "POST":
        branch_actualy = set_branch_session(request)
        try:
            sale = Sale.objects.get(pk=pk)
            form = TypePaymentMethodForm(request.POST)
            if form.is_valid():
                amount = form.cleaned_data['amount']
                payment_method = form.cleaned_data['payment_method']
                mov = create_in_movement(branch_actualy, request.user, payment_method.type_method, form.cleaned_data['description'], amount)

                if not mov:
                    messages.error(request, 'Antes de realizar un pago debe Abrir una Caja.')
                    return redirect('cashregister_app:cashregister_view')

                Payment.objects.create(
                    customer = sale.customer,
                    user_made = request.user,
                    amount = amount,
                    payment_method = payment_method,
                    description = f"Pago de duada de venta Nro: {sale.pk}",
                    sale = sale,
                    movement = mov
                )
                sale.missing_balance -= amount
            if sale.missing_balance <= 0:
                sale.state = "COMPLETADO"
                up_objetives(sale.commision_user, sale)
            sale.save()
            
            messages.success(request, 'Pago realizado con éxito.')
            return redirect('sales_app:sale_detail_view', pk=sale.pk)
        
        except Sale.DoesNotExist:
            messages.error(request, 'Lo sentimos, no pudimos encontrar la Venta.')
            return redirect('sales_app:sales_list_view')
    messages.error(request, 'La petición no es válida.')
    return redirect('sales_app:sale_detail_view', pk=pk)




def export_sale_list_to_excel(request, pk):
    # Para la generacion de excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    
    
    branch_actualy = set_branch_session(request)
    
    list_sales = Sale.objects.filter(branch=branch_actualy)

    if not list_sales:
        raise ValueError('No hay productos para exportar.') # modificar error
    
    # Crear un libro de trabajo de Excel
    workbook = Workbook()
    worksheet = workbook.active

    # Definir estilos personalizados para los encabezados
    header_style = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0b1727', end_color='0b1727', fill_type='solid')

    # Definir los encabezados de las columnas
    exclude_fields_user = ["deleted_at", "receipt", "updated_at", "id", "user_made", "refund_date", "branch"]
    headers = [campo[1] for campo in obtener_nombres_de_campos(Sale, *exclude_fields_user)]

    # Aplicar estilos a los encabezados y escribir los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_style
        cell.fill = header_fill

    # Modificar el ancho de la columna (ajustar segÃºn tus necesidades)
    #################################################
    try: 
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter
    #################################################
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 25

    # Agregar los datos de los empleados a la hoja de cÃ¡lculo
    for row_num, sale in enumerate(list_sales, 2):
        worksheet.cell(row=row_num, column=1, value=sale.customer)
        worksheet.cell(row=row_num, column=2, value=sale.state)
        worksheet.cell(row=row_num, column=3, value=str(sale.discount))
        worksheet.cell(row=row_num, column=4, value=str(sale.discount_extra))
        worksheet.cell(row=row_num, column=5, value=str(sale.missing_balance))
        worksheet.cell(row=row_num, column=6, value=str(sale.subtotal))
        worksheet.cell(row=row_num, column=7, value=str(sale.total))

    # Crear una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lista de ventas - Sucursal %s.xlsx' %(branch_actualy.name)

    workbook.save(response)

    return response